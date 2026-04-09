#!/usr/bin/env python3
import json
import re
import sys
import io
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime, date

YADISK_PUBLIC_KEY = "https://disk.360.yandex.ru/i/pbXk6jtc2OeG5Q"
INDEX_HTML = "index.html"

MONTHS_RU = {
    "января": "01", "февраля": "02", "марта": "03", "апреля": "04",
    "мая": "05", "июня": "06", "июля": "07", "августа": "08",
    "сентября": "09", "октября": "10", "ноября": "11", "декабря": "12"
}


def download_from_yadisk():
    print("Получаю ссылку на скачивание с Яндекс Диска...")
    api_url = "https://cloud-api.yandex.net/v1/disk/public/resources/download?public_key=" + urllib.parse.quote(YADISK_PUBLIC_KEY)
    req = urllib.request.Request(api_url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = json.load(resp)
    download_url = data["href"]
    print("Скачиваю файл...")
    req2 = urllib.request.Request(download_url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req2, timeout=60) as resp:
        content = resp.read()
    print(f"Скачано: {len(content)} байт")
    return content


def parse_act_date(text):
    """Парсит дату из строки вида 'Акт № ЧШ-00000001 от 14 января 2026 г.'"""
    m = re.search(r'от\s+(\d{1,2})\s+(\w+)\s+(\d{4})', text, re.IGNORECASE)
    if m:
        day, month_str, year = m.group(1), m.group(2).lower(), m.group(3)
        month = MONTHS_RU.get(month_str)
        if month:
            return f"{year}-{month}-{day.zfill(2)}"
    return ""



def classify_system(text):
    t = text.lower()
    if any(x in t for x in ['спринклер', 'орошени', 'огнезащит', 'огнестойкост', 'апс', 'аупт', 'аувп', 'ипр', ' пк ', 'пк №', 'впв', 'извещател', 'пожарн', 'эвакуац', 'дымов', 'мезонин', 'лду', 'противопожарн']):
        return 'Пожарная безопасность'
    if any(x in t for x in ['пол ', 'полов', 'топпинг', 'покрыти пол', 'напольн']):
        return 'Покрытие полов'
    if any(x in t for x in ['асфальт']):
        return 'Покрытие полов'
    if any(x in t for x in ['лвж', 'гж', 'гсм', 'аэрозол', 'топлив', 'горюч']):
        return 'ЛВЖ/ГСМ'
    if any(x in t for x in ['ворота', 'дверь', 'двери', 'замок', 'петли', 'докшелтер', 'аппарел', 'доквеллер', 'шлагбаум', 'жалюзи']):
        return 'Ворота и двери'
    if any(x in t for x in ['кабел', 'щит', 'силовой', 'зарядн', 'розетк', 'удлинител', 'электр', 'слаботочн']):
        return 'Электрика'
    if any(x in t for x in ['стена', 'панел', 'перила', 'забор', 'ограждени', 'отбойник', 'бетон', 'конструкци', 'кровл', 'крыш', 'арматур', 'пандус']):
        return 'Конструктив'
    if any(x in t for x in ['вентиляц', 'вытяжк', 'приток', 'воздух']):
        return 'Вентиляция'
    if any(x in t for x in ['проживани', 'абч', 'бытовк', 'санузел', 'парковк', 'курилк', 'насажден']):
        return 'АБЧ/Территория'
    if any(x in t for x in ['стеллаж', 'склад', 'хранени', 'зона']):
        return 'Склад/Хранение'
    if any(x in t for x in ['план эвакуац', 'план', 'документ', 'журнал', 'инструкци', 'ответствен']):
        return 'Документация'
    return 'Прочее'

def parse_remarks(xlsx_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    print(f"=== Листы в файле: {wb.sheetnames} ===")

    # Ищем нужный лист
    ws = None
    for name in wb.sheetnames:
        if "акт" in name.lower() or "претензи" in name.lower() or "сфн" in name.lower():
            ws = wb[name]
            print(f"Используем лист: '{name}'")
            break
    if ws is None:
        ws = wb.active
        print(f"Используем активный лист: '{ws.title}'")

    remarks = []
    current_act_num = None
    current_act_date = None
    current_object = None
    row_id = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row_idx < 315:
            continue

        col0 = str(row[0]).strip() if row[0] else ""

        # Заголовок акта: "Акт № ЧШ-00000001 от 14 января 2026 г."
        is_act_header = bool(re.match(r'^Акт\s+№', col0, re.IGNORECASE))

        if is_act_header:
            # Извлекаем номер акта
            num_match = re.search(r'((?:ЧШ|ДД|чш|дд|[А-Яа-я]{2})-\d+|\d{2}-\d+)', col0)
            current_act_num = num_match.group(1) if num_match else col0[:30]
            current_act_date = parse_act_date(col0)
            # Объект в col1 (Чашниково / Дедовск и т.д.)
            current_object = str(row[1]).strip() if row[1] else ""
            continue

        if not col0 or col0 in ("None", "") or not current_act_num:
            continue
        # Пропускаем строки-заголовки таблицы
        if any(x in col0.lower() for x in ["замечание", "наименование", "пункт", "№ п/п"]):
            continue

        text = col0
        # Объект замечания тоже в col1
        obj = str(row[1]).strip() if len(row) > 1 and row[1] and str(row[1]) != "None" else current_object
        risk_raw = str(row[2]).strip() if len(row) > 2 and row[2] and str(row[2]) != "None" else ""
        deadline1 = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        deadline2 = str(row[7]).strip() if len(row) > 7 and row[7] else ""
        comment = str(row[9]).strip() if len(row) > 9 and row[9] else ""
        comment2 = str(row[10]).strip() if len(row) > 10 and row[10] else ""
        comment3 = str(row[11]).strip() if len(row) > 11 and row[11] else ""
        full_comment = " | ".join(filter(None, [comment, comment2, comment3]))
        responsible = str(row[5]).strip() if len(row) > 5 and row[5] else ""

        # Дедлайн
        deadline = ""
        for d in [deadline2, deadline1]:
            if d and d not in ("None", "") and len(d) >= 8:
                # Если дата как объект datetime
                try:
                    if hasattr(d, 'strftime'):
                        deadline = d.strftime("%Y-%m-%d")
                        break
                except Exception:
                    pass
                deadline = d[:10]
                break

        comment_lower = full_comment.lower()
        today = date.today()

        # Статус
        if any(x in comment_lower for x in ["устранено", "выполнено", "готово", "убрали", "демонтир"]):
            status = "done"
        elif any(x in comment_lower for x in ["направлено", "мониторинг", "регулярн", "письмо"]):
            status = "progress"
        elif re.search(r'2026', full_comment):
            status = "progress"
        else:
            status = "open"

        if status != "done" and deadline:
            try:
                dl = datetime.strptime(deadline[:10], "%Y-%m-%d").date()
                if dl < today:
                    status = "overdue"
            except Exception:
                pass

        text_lower = text.lower()
        risk_lower = risk_raw.lower()
        eviction_risk = ('высел' in risk_lower) or any(x in text_lower for x in [
            "мезонин", "огнезащит", "огнестойкост", "лвж", "гж", "гсм",
            "аэрозольн", "проживание", "система орошени"
        ])
        fine_risk = ('штраф' in risk_lower) or any(x in text_lower for x in [
            "штраф", "нарушен", "предписани", "протокол"
        ]) or eviction_risk
        risk = "eviction" if eviction_risk else ("fine" if fine_risk else "")
        is_rvb = "рвб" in comment_lower or "rvb" in comment_lower

        row_id += 1
        remarks.append({
            "id": row_id,
            "object": obj or current_object,
            "actNum": current_act_num,
            "actDate": current_act_date or "",
            "text": text,
            "risk": risk,
            "responsible": responsible,
            "deadline": deadline,
            "comment": full_comment,
            "status": status,
            "evictionRisk": eviction_risk,
            "fineRisk": fine_risk,
            "resp": "rvb" if is_rvb else "",
            "respName": responsible,
            "isRvb": is_rvb,
            "system": classify_system(text)
        })

    print(f"Распарсено {len(remarks)} замечаний")
    return remarks


def update_index_html(remarks):
    with open(INDEX_HTML, encoding="utf-8") as f:
        html = f.read()
    new_json = json.dumps(remarks, ensure_ascii=False, separators=(",", ":"))
    marker_start = 'const STATIC_DATA = ['
    marker_end = '];'
    start_idx = html.find(marker_start)
    if start_idx == -1:
        print("ОШИБКА: STATIC_DATA не найден в index.html")
        return False
    end_idx = html.find(marker_end, start_idx + len(marker_start))
    if end_idx == -1:
        print("ОШИБКА: конец STATIC_DATA не найден")
        return False
    new_data = f'const STATIC_DATA = {new_json};'
    new_html = html[:start_idx] + new_data + html[end_idx + len(marker_end):]
    with open(INDEX_HTML, "w", encoding="utf-8") as f:
        f.write(new_html)
    print(f"index.html обновлён ({len(remarks)} замечаний)")
    return True


if __name__ == "__main__":
    try:
        xlsx_bytes = download_from_yadisk()
        remarks = parse_remarks(xlsx_bytes)
        if not remarks:
            print("Замечания не найдены — проверьте DEBUG вывод выше")
            sys.exit(1)
        if not update_index_html(remarks):
            sys.exit(1)
        print("Готово!")
    except Exception as e:
        print(f"Ошибка: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
